"""Excel / CSV / PDF disari aktarma.

Kutuphaneler: openpyxl (MIT), reportlab (BSD-3-Clause) - ikisi de ticari
kullanima uygundur.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Sequence
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

BRAND = "#5C1A2B"  # bordo
BRAND_LIGHT = "#F3E7EA"


# ----------------------------------------------------------------- EXCEL
def to_xlsx(
    title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]], *, sheet: str = "Rapor"
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BRAND.lstrip("#"))
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=header)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="8B2942")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border

    for r, row in enumerate(rows, start=4):
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.border = border
            if isinstance(value, (int, float)):
                c.alignment = Alignment(horizontal="right")
                c.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"

    for col, header in enumerate(headers, start=1):
        longest = max(
            [len(str(header))] + [len(str(r[col - 1])) for r in rows if col - 1 < len(r)] or [10]
        )
        ws.column_dimensions[get_column_letter(col)].width = min(46, max(12, longest + 3))

    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------- CSV
def to_csv(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    # Excel'in Turkce karakterleri dogru okumasi icin UTF-8 BOM
    return "﻿".encode() + buf.getvalue().encode("utf-8")


# ------------------------------------------------------------------- PDF
def to_pdf(
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    subtitle: str = "",
    footer: str = "Akıllı Şaraphane Yönetim Sistemi",
) -> bytes:
    buf = io.BytesIO()
    page = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=title,
        author=footer,
    )
    styles = getSampleStyleSheet()
    h = ParagraphStyle(
        "BaslikTR",
        parent=styles["Title"],
        fontSize=16,
        textColor=colors.HexColor(BRAND),
        spaceAfter=4,
    )
    sub = ParagraphStyle(
        "AltBaslik", parent=styles["Normal"], fontSize=9, textColor=colors.grey
    )
    cell_style = ParagraphStyle(
        "Hucre", parent=styles["Normal"], fontSize=7.5, leading=9.5
    )
    head_style = ParagraphStyle(
        "Baslik2",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )

    story: list[Any] = [Paragraph(title, h)]
    if subtitle:
        story.append(Paragraph(subtitle, sub))
    story.append(Spacer(1, 6 * mm))

    data: list[list[Any]] = [[Paragraph(str(x), head_style) for x in headers]]
    for row in rows:
        data.append([Paragraph("" if v is None else str(v), cell_style) for v in row])

    if len(data) == 1:
        data.append([Paragraph("Kayıt bulunamadı.", cell_style)] + [""] * (len(headers) - 1))

    avail = page[0] - 28 * mm
    col_width = avail / max(1, len(headers))
    table = Table(data, colWidths=[col_width] * len(headers), repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9C7CC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(BRAND_LIGHT)]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(footer, sub))

    doc.build(story)
    return buf.getvalue()


def render(
    fmt: str,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    subtitle: str = "",
) -> tuple[bytes, str, str]:
    """(icerik, mime, uzanti) doner."""
    if fmt == "csv":
        return to_csv(headers, rows), "text/csv; charset=utf-8", "csv"
    if fmt == "pdf":
        return (
            to_pdf(title, headers, rows, subtitle=subtitle),
            "application/pdf",
            "pdf",
        )
    return (
        to_xlsx(title, headers, rows),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    )
